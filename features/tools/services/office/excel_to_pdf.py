"""Excel to PDF via LibreOffice (fallback: openpyxl + PyMuPDF)."""
import shutil
import subprocess
import tempfile
from pathlib import Path

import fitz

from features.tools.models import ToolJob
from features.tools.services import file_handler
from features.tools.services.exceptions import ToolError
from features.tools.services.pdf_utils import render_preview


def _find_soffice():
    for cmd in ("soffice", "libreoffice", "/Applications/LibreOffice.app/Contents/MacOS/soffice"):
        if cmd.startswith("/"):
            if Path(cmd).exists():
                return cmd
        else:
            found = shutil.which(cmd)
            if found:
                return found
    return None


def _excel_to_pdf_fallback(src: Path, out: Path, job: ToolJob) -> dict:
    """Convert XLSX to PDF using openpyxl + PyMuPDF text rendering."""
    try:
        import openpyxl
    except ImportError as exc:
        raise ToolError("openpyxl not installed.") from exc

    try:
        wb = openpyxl.load_workbook(str(src), read_only=True, data_only=True)
    except Exception as exc:
        raise ToolError(f"Cannot read Excel file: {exc}") from exc

    doc = fitz.open()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        page = doc.new_page(width=842, height=595)
        y = 40
        page.insert_text(fitz.Point(40, y), f"Sheet: {sheet_name}", fontsize=14, fontname="helv")
        y += 24

        for row in ws.iter_rows(values_only=True):
            if y > 560:
                page = doc.new_page(width=842, height=595)
                y = 40
            row_text = "  |  ".join(str(c) if c is not None else "" for c in row)
            if row_text.strip():
                page.insert_text(fitz.Point(40, y), row_text[:120], fontsize=9, fontname="cour")
                y += 14

    wb.close()
    doc.save(str(out), garbage=3, deflate=True)
    doc.close()

    out_rel = file_handler.to_relative(out)
    return {
        "output_file": out_rel,
        "compression_stat": {"converter": "openpyxl-fallback"},
        "preview_file": render_preview(job, out_rel),
    }


def run(job: ToolJob) -> dict:
    src_rel = job.input_files[0]
    src = file_handler.to_absolute(src_rel)
    out = file_handler.output_path(job, "converted.pdf")
    soffice = _find_soffice()

    if not soffice:
        return _excel_to_pdf_fallback(src, out, job)

    with tempfile.TemporaryDirectory() as tmp:
        result = subprocess.run(
            [soffice, "--headless", "--convert-to", "pdf", "--outdir", tmp, str(src)],
            capture_output=True, text=True, timeout=120,
        )
        if result.returncode != 0:
            raise ToolError("Conversion failed. Ensure the Excel file is valid.")
        pdfs = list(Path(tmp).glob("*.pdf"))
        if not pdfs:
            raise ToolError("No PDF output produced.")
        shutil.copy(pdfs[0], out)

    out_rel = file_handler.to_relative(out)
    return {
        "output_file": out_rel,
        "compression_stat": {"converter": "libreoffice"},
        "preview_file": render_preview(job, out_rel),
    }
